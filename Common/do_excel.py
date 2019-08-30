"""
# @Time    : 2019/8/29 19:32
# @Author  : 谢超
# @File    : do_excel.py
# @Function: 完成excel的读和写
"""


import openpyxl, sys
from Common import contants
from Common.do_log import MyLog


class Case:
    """
    把拥有同一特征的归为一个类
    测试用例类，每个测试用例，实际上就是它的一个实例
    """

    def __init__(self):
        self.case_id = None
        self.case_name = None
        self.method = None
        self.url = None
        self.data = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None


class DoExcel:
    """
    第一种存储数据的方法：
    1.循环遍历excel，获取到的每一行值赋值给Case类的实例中，存储在名为cases的列表中
    2.把所有属性放到一个字典中， 实例.__dict__
    3.和之前把读取到的数据以键值对存在一个列表中的方式一样
    4.这种方式是ddt数据驱动中以类属性来接收数据
    """

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        try:
            self.workbook = openpyxl.load_workbook(self.file_name)
        except Exception as e:
            MyLog(__name__).my_log().error('出错啦！没有这样的文件或目录，或者文件路径不正确。报错：{}'.format(e))
            sys.exit()  # 出错了就停止运行程序
        self.sheet = self.workbook[self.sheet_name]

    def get_cases(self):
        cases = []  # 列表，存放所有的测试用例
        for row in range(2, self.sheet.max_row + 1):
            case = Case()  # 实例
            case.case_id = self.sheet.cell(row=row, column=1).value
            case.case_name = self.sheet.cell(row=row, column=2).value
            case.method = self.sheet.cell(row=row, column=3).value
            case.url = self.sheet.cell(row=row, column=4).value
            case.data = self.sheet.cell(row=row, column=5).value
            case.expected = self.sheet.cell(row=row, column=6).value
            case.sql = self.sheet.cell(row=row, column=9).value
            cases.append(case)
        self.workbook.close()
        return cases

    def write_case(self, row, actual, result):
        sheet = self.workbook[self.sheet_name]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.workbook.save(self.file_name)
        self.workbook.close()


# class Doexcel_2:
#     """
#     第二种存储数据的方法：
#     1.循环遍历excel，把获取到的每一行的值保存在一个base_list列表中，所有列表保存在一个cases大列表中
#     2.这种方式是ddt数据驱动中以变量名来接收数据
#     """
#
#     def __init__(self, file_name, sheet_name):
#         self.file_name = file_name
#         self.sheet_name = sheet_name
#         try:
#             self.workbook = openpyxl.load_workbook(file_name)
#         except:
#             print('出错啦！没有这样的文件或目录')
#         self.sheet = self.workbook[sheet_name]
#
#     def get_cases(self):
#         max_row = self.sheet.max_row  # 获取最大行
#
#         cases = []  # 列表，存放所有的测试用例
#         for row in range(2, max_row + 1):
#             base_list = [self.sheet.cell(row=row, column=1).value,
#                          self.sheet.cell(row=row, column=2).value,
#                          self.sheet.cell(row=row, column=3).value,
#                          self.sheet.cell(row=row, column=4).value,
#                          self.sheet.cell(row=row, column=5).value,
#                          self.sheet.cell(row=row, column=6).value
#                          ]
#             cases.append(base_list)
#         self.workbook.close()
#         return cases
#
#     def write_case(self, row, actual, result):
#         sheet = self.workbook[self.sheet_name]
#         sheet.cell(row, 7).value = actual
#         sheet.cell(row, 8).value = result
#         self.workbook.save(self.file_name)
#         self.workbook.close()
#
#
# class Doexcel_3:
#     """
#     第三种存储数据的方法：
#     1.循环遍历excel，把获取到的每一行的值保存在以键值对保存在字典中，所有字典保存在一个cases大列表中
#     2.这种方式是ddt数据驱动中以key来接收数据
#     """
#
#     def __init__(self, file_name, sheet_name):
#         self.file_name = file_name
#         self.sheet_name = sheet_name
#         try:
#             self.workbook = openpyxl.load_workbook(file_name)
#         except:
#             print('出错啦！没有这样的文件或目录')
#         self.sheet = self.workbook[sheet_name]
#
#     def get_cases(self):
#         max_row = self.sheet.max_row  # 获取最大行
#
#         cases = []  # 列表，存放所有的测试用例
#         for row in range(2, max_row + 1):
#             case = {}
#             case['case_id'] = self.sheet.cell(row=row, column=1).value
#             case['case_name'] = self.sheet.cell(row=row, column=2).value
#             case['method'] = self.sheet.cell(row=row, column=3).value
#             case['url'] = self.sheet.cell(row=row, column=4).value
#             case['data'] = self.sheet.cell(row=row, column=5).value
#             case['expected'] = self.sheet.cell(row=row, column=6).value
#
#             cases.append(case)
#         self.workbook.close()
#         return cases
#
#     def write_case(self, row, actual, result):
#         sheet = self.workbook[self.sheet_name]
#         sheet.cell(row, 7).value = actual
#         sheet.cell(row, 8).value = result
#         self.workbook.save(self.file_name)
#         self.workbook.close()


if __name__ == '__main__':
    a=DoExcel(contants.case_file,"GetBlackBlackList")
    b=a.get_cases()
    print(b)